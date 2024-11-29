import openpyxl
import pandas as pd
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta, timezone

# Create a new workbook and select the active sheet
wb = openpyxl.Workbook()
ws = wb.active

# Starting date and initial row setup
start_date = pd.Timestamp('2018-08-1')
rows_per_date = 5
current_date = start_date

# Generate the pattern of repeating dates in column A
row = 2  # Start at row 2 (A2)
while row < 2000:  # Limit to 100 rows for this example
    for _ in range(rows_per_date):
        ws[f'A{row}'] = current_date.strftime('%Y-%m-%d')
        row += 1
    current_date += timedelta(days=31)  # Increase the date by 7 days

# Save the generated Excel file
output_excel_path = 'C:\\Users\\nmercado\\Documents\\Scripts\\fechas.xlsx'
wb.save(output_excel_path)

output_excel_path